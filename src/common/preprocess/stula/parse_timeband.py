"""
社会生活基本調査 (令和3年, 調査票A) 公表「時間帯編」Excel表のパーサ

データ入手 (e-Stat 統計コード00200533 / 時間帯編47表, curl はブラウザ相当UA+Refererで可):
    data/raw/opened/STULA2021/<statInfId>.xlsx に配置済みの表を処理する。
        000032224341: 曜日,男女,就業状態,行動の種類,年齢,時刻区分別行動者率(15歳以上) - 平日
        000032224342: 同 - 土曜日
        000032224343: 同 - 日曜日
        000032224339: 曜日,行動の種類,男女,時刻区分別行動者率(15歳以上) - 全国,都道府県

表の構造 (000032224341 で実検証):
    行1-2: タイトル / 行5: ブロック見出し(推定人口/行動者率) / 行7: 時刻区分ラベル
    行9: キー列名 (曜日/地域区分/男女/ふだんの就業状態/行動の種類/年齢)
    行10-: データ。キー値は "コード_ラベル" 形式 (例 '1_平日', '01_睡眠', '03_25～29歳')
    時刻区分: 01_0:00-0:15 .. 96_23:45-24:00 の96列 (%表記)。総数行の率は '-'
    行動: 20分類 + 00_総数 + R1-R3_(再掲)1〜3次活動
    年齢: 5歳階級16区分 (01_15～19歳 .. 15_85歳以上) + 総数 + 再掲R1-R6

処理:
    1. 行9からキー列、行7から時刻96列を自動検出 (表ごとのキー構成差を吸収)
    2. 0:00起点の96列を **04:00起点へ回転** (ATUS前処理の diary day と整合):
       出力 s_j = 原表の時刻列 (j+16) mod 96  (s0 = 04:00-04:15)
    3. '-' は NaN。率は % のまま保持 (0..100)
    4. 検証: 行動01..20の率を層内で合計すると各スロット ≈100% (主行動は排他的)

出力: data/processed/stula/<name>.csv (wide形式: キー列 + population_k + s0..s95)
"""

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[4]   # src/common/preprocess/stula -> repo root
RAW_DIR   = REPO_ROOT / "data" / "raw" / "opened" / "STULA2021"
OUT_DIR   = REPO_ROOT / "data" / "processed" / "stula"

# 処理対象: statInfId -> 出力名
TABLES = {
    "000032224341": "timeband_weekday",
    "000032224342": "timeband_saturday",
    "000032224343": "timeband_sunday",
    "000032224339": "timeband_prefecture",
}

N_SLOTS   = 96
ROT       = 16   # 0:00起点 -> 04:00起点の回転量 (4h / 15min)
HEADER_KEYS_ROW = 9   # キー列名の行
DATA_START_ROW  = 10
SLOT_LABEL_ROW  = 7   # 時刻区分ラベルの行

# キー列名の正規化 (表側の日本語ヘッダ -> 出力列名)
KEY_RENAME = {
    "曜日": "daytype", "地域区分": "region", "男女": "gender",
    "ふだんの就業状態": "employment", "行動の種類": "activity", "年齢": "age_class",
}


def parse_table(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    grid_head = [next(rows) for _ in range(DATA_START_ROW - 1)]

    # キー列: 行9で非空のセル / 時刻列: 行7で "NN_H:MM - H:MM" 形式のセル
    key_cols = {c: str(v).strip() for c, v in enumerate(grid_head[HEADER_KEYS_ROW - 1])
                if v is not None and str(v).strip() != ""}
    slot_cols = {}
    for c, v in enumerate(grid_head[SLOT_LABEL_ROW - 1]):
        if v and "_" in str(v) and ":" in str(v):
            slot_cols[c] = int(str(v).split("_")[0]) - 1   # 0-based 0:00起点スロット
    assert len(slot_cols) == N_SLOTS, f"時刻区分が96列でない: {len(slot_cols)} ({path.name})"
    # 推定人口列: キー列と時刻列の間の残り1列 (行8の単位=千人)
    pop_col = [c for c in range(min(slot_cols)) if c not in key_cols][-1]

    def to_rate(v):
        # 欠損マーカーは複数種 ('-', '…', '...', 'X' 等) → 数値化できないものは一律 NaN
        try:
            return float(v)
        except (TypeError, ValueError):
            return np.nan

    records = []
    for row in rows:
        if row[list(key_cols)[0]] is None:
            continue
        rec = {KEY_RENAME.get(name, name): str(row[c]).strip() for c, name in key_cols.items()}
        rec["population_k"] = to_rate(row[pop_col])
        vals = np.full(N_SLOTS, np.nan)
        for c, k0 in slot_cols.items():
            vals[(k0 - ROT) % N_SLOTS] = to_rate(row[c])   # 04:00起点へ回転
        for j in range(N_SLOTS):
            rec[f"s{j}"] = vals[j]
        records.append(rec)
    return pd.DataFrame(records)


def validate(df: pd.DataFrame, name: str):
    """主行動の排他性: 行動01..20を層内合計すると各スロット≈100%"""
    if "activity" not in df.columns:
        print(f"  [{name}] activity列なし: 合計検証スキップ")
        return
    act20 = df[df["activity"].str.match(r"^(0[1-9]|1\d|20)_")]
    keys = [c for c in df.columns if c not in ("activity", "population_k") and not c.startswith("s")]
    slot_cols = [f"s{j}" for j in range(N_SLOTS)]
    totals = act20.groupby(keys, observed=True)[slot_cols].sum()
    # 率が公表されていない層 (全て NaN → 合計0) は除外して評価
    nonzero = totals[totals.sum(axis=1) > 0]
    lo, hi = nonzero.to_numpy().min(), nonzero.to_numpy().max()
    print(f"  [{name}] 20分類スロット合計: min={lo:.1f}% max={hi:.1f}% (層数 {len(nonzero)})")
    assert 95.0 < lo and hi < 105.0, "主行動の合計が100%から大きく外れる層がある"


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for stat_id, name in TABLES.items():
        path = RAW_DIR / f"{stat_id}.xlsx"
        if not path.exists():
            print(f"[skip] {path} が無い (e-Stat statInfId={stat_id} をDLして配置)")
            continue
        print(f"parsing {path.name} ({name}) ...")
        df = parse_table(path)
        key_desc = {c: df[c].nunique() for c in df.columns if not c.startswith("s") and c != "population_k"}
        print(f"  rows={len(df)}  keys={key_desc}")
        validate(df, name)
        out = OUT_DIR / f"{name}.csv"
        df.to_csv(out, index=False)
        print(f"  -> {out}")


if __name__ == "__main__":
    main()
